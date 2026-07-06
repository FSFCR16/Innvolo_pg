# -*- coding: utf-8 -*-
"""
FASE 0a — Volcar el Excel a WooCommerce (DATOS, sin imágenes).
- Lee precios/descripcion/materiales/colores/tecnica del Excel (fuente de verdad).
- Crea 4 productos nuevos (Camisa, Camiseta, Overol, Botella Metalica).
- Renombra Polo (13) y Bata (215).
- Escribe meta precio_20/50/100/500, tecnica_recomendada, materiales, colores_disponibles.
- Pone regular_price = precio_20 en las variaciones (para que Woo muestre "desde").
- Idempotente: los nuevos se saltan si ya existe su SKU.
- Backup de los productos existentes antes de escribir.

USO:
    python woo_fase0.py            # DRY-RUN (no escribe nada, muestra tabla)
    python woo_fase0.py --write    # ESCRIBE en WooCommerce
"""
import os, sys, json, time
sys.stdout.reconfigure(encoding="utf-8")
from dotenv import load_dotenv
import requests
from requests.auth import HTTPBasicAuth
import urllib3
from openpyxl import load_workbook

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
load_dotenv()

BASE = os.getenv("WC_BASE_URL") + "/wp-json/wc/v3"
AUTH = HTTPBasicAuth(os.getenv("WC_KEY"), os.getenv("WC_SECRET"))
EXCEL = r"C:\Users\Netapplicatiosn\Downloads\Catalogo Precios INNVOLO.xlsx"
IMG_DIR = r"C:\Users\Netapplicatiosn\Downloads\INNVOLO\IMAGENES CATALOGO"

WRITE = "--write" in sys.argv
SLEEP = 0.15

# IDs de atributos globales (de woo_catalog_state.json)
ATTR = {"color": 1, "talla": 2, "material": 3, "tecnica_impresion": 4, "capacidad": 5}
TALLAS_BASE = ["S", "M", "L", "XL"]
COLORES_BASE = ["Negro", "Blanco", "Azul marino", "Gris"]

# Mapeo por # del Excel (1..36):
#   woo_id  -> producto existente a actualizar (None = crear nuevo)
#   cat     -> categoria para los nuevos
#   img     -> token HHMMSS del archivo de imagen (para 0b; aqui solo se valida)
#   tipo3d  -> meta tipo_3d para nuevos
#   sku     -> sku para nuevos
MAPEO = {
    1:  dict(woo_id=None, cat=70, img="090026", tipo3d="torso", sku="CAMISA-CORP"),
    2:  dict(woo_id=None, cat=70, img="195649", tipo3d="torso", sku="CAMISETA-CORP"),
    3:  dict(woo_id=13,   img="194441"),
    4:  dict(woo_id=35,   img="231938"),
    5:  dict(woo_id=30,   img="232344"),
    6:  dict(woo_id=57,   img="200231"),
    7:  dict(woo_id=40,   img="200530"),
    8:  dict(woo_id=74,   img="201645"),
    9:  dict(woo_id=91,   img="184438"),
    10: dict(woo_id=108,  img="184400"),
    11: dict(woo_id=125,  img="200012"),
    12: dict(woo_id=142,  img="200854"),
    13: dict(woo_id=159,  img="202245"),
    14: dict(woo_id=176,  img="202631"),
    15: dict(woo_id=193,  img="230521"),
    16: dict(woo_id=210,  img="233533"),
    17: dict(woo_id=215,  img="040904"),
    18: dict(woo_id=None, cat=85, img="041723", tipo3d="torso", sku="OVEROL-IND"),
    19: dict(woo_id=232,  img="235549"),
    20: dict(woo_id=None, cat=88, img="235549", tipo3d="recipiente", sku="BOTELLA-MET"),
    21: dict(woo_id=245,  img="235937"),
    22: dict(woo_id=250,  img="235937"),
    23: dict(woo_id=255,  img="235216"),
    24: dict(woo_id=260,  img="003711"),
    25: dict(woo_id=265,  img="190802"),
    26: dict(woo_id=270,  img="190747"),
    27: dict(woo_id=275,  img="003729"),
    28: dict(woo_id=280,  img="002523"),
    29: dict(woo_id=285,  img="002911"),
    30: dict(woo_id=290,  img="182349"),
    31: dict(woo_id=307,  img="053510"),
    32: dict(woo_id=324,  img="011315"),
    33: dict(woo_id=329,  img="010724"),
    34: dict(woo_id=339,  img="005602"),
    35: dict(woo_id=334,  img="005833"),
    36: dict(woo_id=344,  img="011155"),
}


def round10(x):
    return int(round(float(x) / 10.0)) * 10


def leer_excel():
    wb = load_workbook(EXCEL, data_only=True)
    ws = wb["Catálogo y Precios"]
    rows = list(ws.iter_rows(values_only=True))
    prods, n = {}, 0
    for r in rows[2:]:
        if r[2] is None or r[7] is None:   # sin nombre o sin precio -> seccion/blank
            continue
        n += 1
        prods[n] = dict(
            nombre=str(r[2]).strip(),
            descripcion=str(r[3]).strip() if r[3] else "",
            materiales=str(r[4]).strip() if r[4] else "",
            colores=str(r[5]).strip() if r[5] else "",
            p20=round10(r[7]), p50=round10(r[8]), p100=round10(r[9]), p500=round10(r[10]),
            tecnica=str(r[11]).strip() if r[11] else "",
        )
    return prods


def resolver_imagenes():
    idx = {}
    for f in os.listdir(IMG_DIR):
        if f.lower().endswith(".png"):
            parts = f.split("_")
            if len(parts) >= 3:
                idx[parts[2]] = os.path.join(IMG_DIR, f)
    return idx


def req(method, url, **kw):
    kw.setdefault("verify", False)
    kw.setdefault("timeout", 60)
    r = requests.request(method, url, auth=AUTH, **kw)
    if r.status_code >= 400:
        print(f"    ! HTTP {r.status_code} {method} {url.split('/wc/v3')[-1]} -> {r.text[:160]}")
        return None
    return r.json()


def meta_de(p):
    return {"key": "materiales",           "value": p["materiales"]}, \
           {"key": "colores_disponibles",  "value": p["colores"]}, \
           {"key": "tecnica_recomendada",  "value": p["tecnica"]}, \
           {"key": "precio_20",            "value": str(p["p20"])}, \
           {"key": "precio_50",            "value": str(p["p50"])}, \
           {"key": "precio_100",           "value": str(p["p100"])}, \
           {"key": "precio_500",           "value": str(p["p500"])}


def attrs_nuevos(tipo3d):
    if tipo3d == "recipiente":
        return [
            {"id": ATTR["material"], "position": 0, "visible": True, "variation": False, "options": ["Acero inoxidable"]},
            {"id": ATTR["color"], "position": 1, "visible": True, "variation": True, "options": COLORES_BASE},
        ], [{"color": c} for c in COLORES_BASE]
    return [
        {"id": ATTR["material"], "position": 0, "visible": True, "variation": False, "options": ["Algodón"]},
        {"id": ATTR["talla"], "position": 1, "visible": True, "variation": True, "options": TALLAS_BASE},
        {"id": ATTR["color"], "position": 2, "visible": True, "variation": True, "options": COLORES_BASE},
    ], [{"talla": t, "color": c} for t in TALLAS_BASE for c in COLORES_BASE]


def main():
    prods = leer_excel()
    imgs = resolver_imagenes()
    assert len(prods) == 36, f"Excel dio {len(prods)} productos (esperaba 36)"

    modo = "ESCRITURA (--write)" if WRITE else "DRY-RUN (no escribe)"
    print(f"\n{'='*100}\n  FASE 0a — {modo}\n{'='*100}")
    print(f"  {'#':>2} {'PRODUCTO':<28} {'ACCION':<11} {'ID':>5}  {'P20':>7} {'P50':>7} {'P100':>7} {'P500':>7}  {'TECNICA':<20} IMG")
    print("  " + "-"*98)
    plan = []
    for n in range(1, 37):
        p = prods[n]; m = MAPEO[n]
        nuevo = m["woo_id"] is None
        accion = "CREAR" if nuevo else ("RENOMBRAR" if n in (3, 17) else "ACTUALIZAR")
        img_ok = "OK" if m["img"] in imgs else "FALTA!"
        print(f"  {n:>2} {p['nombre'][:28]:<28} {accion:<11} {str(m['woo_id'] or '-'):>5}  "
              f"{p['p20']:>7} {p['p50']:>7} {p['p100']:>7} {p['p500']:>7}  {p['tecnica'][:20]:<20} {m['img']}:{img_ok}")
        plan.append((n, p, m, nuevo))
    print("  " + "-"*98)
    faltan = [m["img"] for _, _, m, _ in plan if m["img"] not in imgs]
    print(f"  Imagenes encontradas: {36-len(faltan)}/36" + ("  FALTAN: "+",".join(faltan) if faltan else "  (todas)"))

    if not WRITE:
        print("\n  >> DRY-RUN. Nada escrito. Corre con --write para aplicar.\n")
        return

    # ---- BACKUP de existentes ----
    print("\n  Backup de productos existentes...")
    backup = {}
    for n, p, m, nuevo in plan:
        if not nuevo:
            d = req("GET", f"{BASE}/products/{m['woo_id']}")
            if d:
                backup[m["woo_id"]] = {"name": d["name"], "description": d["description"],
                                        "short_description": d["short_description"],
                                        "meta_data": d["meta_data"]}
    with open("woo_fase0_backup.json", "w", encoding="utf-8") as f:
        json.dump(backup, f, ensure_ascii=False, indent=2)
    print(f"  Backup guardado: woo_fase0_backup.json ({len(backup)} productos)")

    # ---- ESCRITURA ----
    creados = actualizados = 0
    id_map = {}
    for n, p, m, nuevo in plan:
        meta = list(meta_de(p))
        desc = f"<p>{p['descripcion']}</p>" if p["descripcion"] else ""
        if nuevo:
            # idempotencia por SKU
            ex = req("GET", f"{BASE}/products", params={"sku": m["sku"]})
            if ex:
                print(f"  = #{n} {p['nombre']}: ya existe SKU {m['sku']} (id {ex[0]['id']}), solo actualizo")
                pid = ex[0]["id"]
                req("PUT", f"{BASE}/products/{pid}", json={"description": desc, "short_description": p["descripcion"], "meta_data": meta})
                id_map[n] = pid
                actualizados += 1
                time.sleep(SLEEP); continue
            attrs, combos = attrs_nuevos(m.get("tipo3d"))
            payload = {
                "name": p["nombre"], "type": "variable", "status": "publish",
                "catalog_visibility": "visible", "sku": m["sku"],
                "description": desc, "short_description": p["descripcion"],
                "categories": [{"id": m["cat"]}], "attributes": attrs,
                "meta_data": meta + [
                    {"key": "model_3d_url", "value": ""},
                    {"key": "customization_zones", "value": "[]"},
                    {"key": "tipo_3d", "value": m.get("tipo3d", "torso")},
                ],
            }
            d = req("POST", f"{BASE}/products", json=payload)
            if not d:
                print(f"  ! #{n} {p['nombre']}: no creado"); continue
            pid = d["id"]; id_map[n] = pid
            # variaciones (batch) con precio = p20
            vbatch = {"create": []}
            for i, combo in enumerate(combos, 1):
                vattrs = []
                for slug, val in combo.items():
                    vattrs.append({"id": ATTR[slug], "name": slug.capitalize(), "option": val})
                vbatch["create"].append({"sku": f"{m['sku']}-{i:03d}", "regular_price": str(p["p20"]),
                                          "attributes": vattrs, "manage_stock": False})
            req("POST", f"{BASE}/products/{pid}/variations/batch", json=vbatch)
            print(f"  + #{n} {p['nombre']}: CREADO id {pid} (+{len(combos)} variaciones @ {p['p20']})")
            creados += 1
        else:
            pid = m["woo_id"]; id_map[n] = pid
            body = {"description": desc, "short_description": p["descripcion"], "meta_data": meta}
            if n in (3, 17):
                body["name"] = p["nombre"]
            d = req("PUT", f"{BASE}/products/{pid}", json=body)
            # precio en variaciones existentes
            vs = req("GET", f"{BASE}/products/{pid}/variations", params={"per_page": 100}) or []
            if vs:
                upd = {"update": [{"id": v["id"], "regular_price": str(p["p20"])} for v in vs]}
                req("POST", f"{BASE}/products/{pid}/variations/batch", json=upd)
            tag = "RENOMBRADO" if n in (3, 17) else "ACTUALIZADO"
            print(f"  ~ #{n} {p['nombre']}: {tag} id {pid} ({len(vs)} var @ {p['p20']})")
            actualizados += 1
        time.sleep(SLEEP)

    with open("woo_fase0_idmap.json", "w", encoding="utf-8") as f:
        json.dump(id_map, f, ensure_ascii=False, indent=2)
    print(f"\n  LISTO. Creados: {creados} | Actualizados: {actualizados}")
    print(f"  id_map guardado en woo_fase0_idmap.json (para 0b imagenes).")


if __name__ == "__main__":
    main()
